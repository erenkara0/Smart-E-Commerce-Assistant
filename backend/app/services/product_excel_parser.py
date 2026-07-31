from pathlib import Path
from typing import BinaryIO
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import ValidationError

from app.schemas.product_import import (
    ExcelProductRow,
    ProductExcelParseResult,
    ProductImportValidationError,
)


PRODUCTS_WORKSHEET_NAME = "Products"

EXPECTED_COLUMNS = (
    "product_id",
    "sku",
    "name",
    "category",
    "brand",
    "price",
    "currency",
    "description",
    "features",
    "stock",
    "rating",
    "image_url",
    "product_url",
    "is_active",
)

REQUIRED_COLUMNS = frozenset(
    {
        "product_id",
        "sku",
        "name",
        "category",
        "brand",
        "price",
        "currency",
        "description",
        "features",
        "stock",
        "rating",
        "is_active",
    }
)


class ProductExcelParserError(ValueError):
    def __init__(
        self,
        code: str,
        message: str,
    ) -> None:
        super().__init__(message)
        self.code = code


def normalize_header(value: object) -> str:
    if value is None:
        return ""

    return str(value).strip()


def is_empty_row(values: tuple[object, ...]) -> bool:
    return all(
        value is None
        or (
            isinstance(value, str)
            and not value.strip()
        )
        for value in values
    )


def build_header_map(
    headers: tuple[object, ...],
) -> dict[str, int]:
    header_map: dict[str, int] = {}
    duplicate_headers: list[str] = []

    for column_index, raw_header in enumerate(headers):
        header = normalize_header(raw_header)

        if not header:
            continue

        if header in header_map:
            duplicate_headers.append(header)
            continue

        header_map[header] = column_index

    if duplicate_headers:
        duplicate_text = ", ".join(
            sorted(set(duplicate_headers))
        )
        raise ProductExcelParserError(
            code="duplicate_columns",
            message=(
                "The Products worksheet contains duplicate columns: "
                f"{duplicate_text}."
            ),
        )

    missing_columns = sorted(
        REQUIRED_COLUMNS.difference(header_map)
    )

    if missing_columns:
        missing_text = ", ".join(missing_columns)
        raise ProductExcelParserError(
            code="missing_columns",
            message=(
                "The Products worksheet is missing required columns: "
                f"{missing_text}."
            ),
        )

    return header_map


def build_row_payload(
    row_values: tuple[object, ...],
    header_map: dict[str, int],
) -> dict[str, object]:
    payload: dict[str, object] = {}

    for column_name in EXPECTED_COLUMNS:
        column_index = header_map.get(column_name)

        if column_index is None:
            payload[column_name] = None
            continue

        if column_index >= len(row_values):
            payload[column_name] = None
            continue

        payload[column_name] = row_values[column_index]

    return payload


def build_validation_errors(
    excel_row_number: int,
    validation_error: ValidationError,
) -> list[ProductImportValidationError]:
    errors: list[ProductImportValidationError] = []

    for error in validation_error.errors():
        location = error.get("loc", ())
        field = str(location[0]) if location else None
        message = str(
            error.get(
                "msg",
                "Invalid product value.",
            )
        )

        errors.append(
            ProductImportValidationError(
                row=excel_row_number,
                field=field,
                message=message,
            )
        )

    return errors


def parse_product_excel(
    source: str | Path | BinaryIO,
) -> ProductExcelParseResult:
    try:
        workbook = load_workbook(
            filename=source,
            read_only=True,
            data_only=True,
        )
    except FileNotFoundError as exc:
        raise ProductExcelParserError(
            code="file_not_found",
            message="The Excel file could not be found.",
        ) from exc
    except PermissionError as exc:
        raise ProductExcelParserError(
            code="file_access_denied",
            message="The Excel file could not be accessed.",
        ) from exc
    except (InvalidFileException, BadZipFile) as exc:
        raise ProductExcelParserError(
            code="invalid_excel_file",
            message=(
                "The uploaded file is not a valid .xlsx workbook."
            ),
        ) from exc
    except OSError as exc:
        raise ProductExcelParserError(
            code="excel_read_failed",
            message="The Excel file could not be read.",
        ) from exc

    try:
        if PRODUCTS_WORKSHEET_NAME not in workbook.sheetnames:
            raise ProductExcelParserError(
                code="missing_products_worksheet",
                message=(
                    "The workbook must contain a "
                    "'Products' worksheet."
                ),
            )

        worksheet = workbook[PRODUCTS_WORKSHEET_NAME]
        rows = worksheet.iter_rows(values_only=True)

        header_values = next(rows, None)

        if header_values is None or is_empty_row(
            tuple(header_values)
        ):
            raise ProductExcelParserError(
                code="missing_header_row",
                message=(
                    "The Products worksheet must contain "
                    "a header row."
                ),
            )

        header_map = build_header_map(
            tuple(header_values)
        )

        products: list[ExcelProductRow] = []
        errors: list[ProductImportValidationError] = []

        seen_product_ids: dict[str, int] = {}
        seen_skus: dict[str, int] = {}
        total_rows = 0

        for excel_row_number, row_values in enumerate(
            rows,
            start=2,
        ):
            normalized_row = tuple(row_values)

            if is_empty_row(normalized_row):
                continue

            total_rows += 1

            payload = build_row_payload(
                normalized_row,
                header_map,
            )

            try:
                product = ExcelProductRow.model_validate(
                    payload
                )
            except ValidationError as exc:
                errors.extend(
                    build_validation_errors(
                        excel_row_number,
                        exc,
                    )
                )
                continue

            duplicate_errors: list[
                ProductImportValidationError
            ] = []

            first_product_id_row = seen_product_ids.get(
                product.product_id
            )

            if first_product_id_row is not None:
                duplicate_errors.append(
                    ProductImportValidationError(
                        row=excel_row_number,
                        field="product_id",
                        message=(
                            "Duplicate product_id. "
                            "The same value was first used "
                            f"on row {first_product_id_row}."
                        ),
                    )
                )

            first_sku_row = seen_skus.get(product.sku)

            if first_sku_row is not None:
                duplicate_errors.append(
                    ProductImportValidationError(
                        row=excel_row_number,
                        field="sku",
                        message=(
                            "Duplicate SKU. "
                            "The same value was first used "
                            f"on row {first_sku_row}."
                        ),
                    )
                )

            if duplicate_errors:
                errors.extend(duplicate_errors)
                continue

            seen_product_ids[product.product_id] = (
                excel_row_number
            )
            seen_skus[product.sku] = excel_row_number
            products.append(product)

        failed_row_numbers = {
            error.row
            for error in errors
        }

        return ProductExcelParseResult(
            products=products,
            total_rows=total_rows,
            valid_rows=len(products),
            failed_rows=len(failed_row_numbers),
            errors=errors,
        )
    finally:
        workbook.close()